# ═══════════════════════════════════════════════════════════════════════════════
#  NEAM Stock Processor — Streamlit Web App  v2.2
# ═══════════════════════════════════════════════════════════════════════════════

import os, io, datetime
from collections import defaultdict

import streamlit as st
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

APP_NAME   = "NEAM Stock Processor"
APP_VER    = "v2.2"
FOOTER_TXT = "Neam"

# ═══════════════════════════════════════════════════════════════════════════════
#  ١. منطق المعالجة — بدون تغيير
# ═══════════════════════════════════════════════════════════════════════════════

def parse_stock_bytes(file_bytes, sloc_label):
    for enc in ['utf-8', 'windows-1256', 'latin-1', 'utf-8-sig']:
        try:
            soup = BeautifulSoup(file_bytes.decode(enc), 'html.parser')
            break
        except:
            continue
    else:
        return []
    stock = []
    for row in soup.find_all('tr'):
        cells = row.find_all('td')
        if len(cells) < 4:
            continue
        texts = [c.get_text(strip=True).replace('\xa0', ' ').strip() for c in cells]
        mat = texts[0].strip()
        if not mat or mat in ('Material', '*'):
            continue
        desc  = texts[1].strip()
        batch = texts[2].strip()
        try:
            qty = float(texts[3].strip().replace('.', '').replace(',', ''))
        except:
            continue
        stock.append({'sloc': sloc_label, 'material': mat,
                      'description': desc, 'batch': batch, 'unrestricted': qty})
    return stock


def parse_issues_bytes(file_bytes):
    for enc in ['utf-8', 'windows-1256', 'latin-1', 'utf-8-sig']:
        try:
            soup = BeautifulSoup(file_bytes.decode(enc), 'html.parser')
            break
        except:
            continue
    else:
        return defaultdict(float), defaultdict(float), {}
    agg, agg_at, at_meta = defaultdict(float), defaultdict(float), {}
    for row in soup.find_all('tr'):
        cells = row.find_all('td')
        if len(cells) < 8:
            continue
        texts = [c.get_text(strip=True).replace('\xa0', ' ').strip() for c in cells]
        sloc, mat = texts[7].strip(), texts[1].strip()
        if not mat or mat == '*':
            continue
        desc, batch = texts[2].strip(), texts[5].strip()
        try:
            qty = float(texts[3].replace(',', '.').strip())
        except:
            continue
        if desc.startswith('@') and batch == '':
            key = (sloc, mat, desc)
            agg_at[key] += qty
            at_meta[key] = {'sloc': sloc, 'material': mat, 'description': desc}
        else:
            agg[(sloc, mat, desc, batch)] += qty
    return agg, agg_at, at_meta


def get_issued(stock_row, issues_agg):
    sloc, mat, desc, batch = (stock_row['sloc'], stock_row['material'],
                               stock_row['description'], stock_row['batch'])
    key1 = (sloc, mat, desc, batch)
    if key1 in issues_agg and issues_agg[key1] > 0:
        return issues_agg[key1], 1
    total = sum(v for (s, m, d, b), v in issues_agg.items()
                if s == sloc and m == mat and b == batch and b != '')
    if total > 0:
        return total, 2
    total = sum(v for (s, m, d, b), v in issues_agg.items()
                if s == sloc and m == mat and b == '')
    if total > 0:
        return total, 3
    return 0.0, 0


def process(stock_list, issues_agg):
    results = []
    for row in stock_list:
        issued, level = get_issued(row, issues_agg)
        results.append({
            'sloc': row['sloc'], 'material': row['material'],
            'description': row['description'], 'batch': row['batch'],
            'stock': row['unrestricted'], 'issued': issued,
            'remaining': row['unrestricted'] - issued, 'level': level
        })
    return results


def get_unmatched(issues_agg, known_slocs):
    unmatched = defaultdict(list)
    for (sloc, mat, desc, batch), qty in issues_agg.items():
        if sloc not in known_slocs and qty > 0:
            unmatched[sloc].append({'sloc': sloc, 'material': mat,
                                    'description': desc, 'batch': batch,
                                    'issued': qty})
    return unmatched


# ═══════════════════════════════════════════════════════════════════════════════
#  ٢. فلترة الأعمدة — بدون تغيير
# ═══════════════════════════════════════════════════════════════════════════════

ALL_STOCK_HEADERS = ["SLoc","Material","Description","Batch",
                     "Stock","Issued","Remaining","Match Level"]
ALL_STOCK_WIDTHS  = [8, 14, 42, 16, 12, 12, 14, 14]


def filter_headers(show_stock, show_issued, show_level):
    hide = set()
    if not show_stock:  hide.add("Stock")
    if not show_issued: hide.add("Issued")
    if not show_level:  hide.add("Match Level")
    headers, widths = [], []
    for h, w in zip(ALL_STOCK_HEADERS, ALL_STOCK_WIDTHS):
        if h not in hide:
            headers.append(h)
            widths.append(w)
    return headers, widths


def row_to_values(r, show_stock, show_issued, show_level):
    vals = [r['sloc'], r['material'], r['description'], r['batch']]
    if show_stock:  vals.append(r['stock'])
    if show_issued: vals.append(r['issued'])
    vals.append(r['remaining'])
    if show_level:  vals.append(f"L{r['level']}" if r['level'] > 0 else "—")
    return vals


# ═══════════════════════════════════════════════════════════════════════════════
#  ٣. كتابة Excel — بدون تغيير
# ═══════════════════════════════════════════════════════════════════════════════

THIN       = Side(style='thin',   color="B0C4D8")
MED        = Side(style='medium', color="1F4E79")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HDR = Border(left=MED,  right=MED,  top=MED,  bottom=MED)
LEFT_AL    = Alignment(horizontal="left",   vertical="center")
CTR_AL     = Alignment(horizontal="center", vertical="center")


def _xl_info_header(ws, sheet_name, source_files, total_cols, now):
    ws.insert_rows(1, 3)
    c = ws.cell(row=1, column=1, value=f"  {APP_NAME}  —  {sheet_name}")
    c.font      = Font(bold=True, name="Calibri", size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1F4E79")
    c.alignment = LEFT_AL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 26
    meta = (f"  Date: {now.strftime('%Y-%m-%d')}   |   "
            f"Time: {now.strftime('%H:%M:%S')}   |   "
            f"Source: {', '.join(source_files)}")
    m = ws.cell(row=2, column=1, value=meta)
    m.font      = Font(name="Calibri", size=9, italic=True, color="1F4E79")
    m.fill      = PatternFill("solid", fgColor="EBF3FA")
    m.alignment = LEFT_AL
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 5
    for col in range(1, total_cols + 1):
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor="1F4E79")


def _xl_footer(ws, last_row, total_cols):
    r = last_row + 1
    c = ws.cell(row=r, column=1, value=f"  {FOOTER_TXT}")
    c.font      = Font(name="Calibri", size=9, italic=True, color="808080")
    c.alignment = LEFT_AL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)


def write_stock_sheet(ws, results, row_color, sheet_name,
                      source_files, now, show_stock, show_issued, show_level):
    headers, widths = filter_headers(show_stock, show_issued, show_level)
    tc      = len(headers)
    rem_col = headers.index("Remaining") + 1
    _xl_info_header(ws, sheet_name, source_files, tc, now)
    hdr = 4
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=hdr, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = CTR_AL
        cell.border    = BORDER_HDR
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[hdr].height = 22
    FA = PatternFill("solid", fgColor=row_color)
    FB = PatternFill("solid", fgColor="F0F7FF")
    FI = PatternFill("solid", fgColor="FFF3CD")
    num_cols = {rem_col}
    if show_stock  and "Stock"  in headers: num_cols.add(headers.index("Stock")  + 1)
    if show_issued and "Issued" in headers: num_cols.add(headers.index("Issued") + 1)
    for r_idx, row in enumerate(results, hdr + 1):
        issued    = row['issued']
        remaining = row['remaining']
        fill      = FI if issued > 0 else (FA if r_idx % 2 == 0 else FB)
        values    = row_to_values(row, show_stock, show_issued, show_level)
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = LEFT_AL if col == 3 else CTR_AL
            cell.border    = BORDER
            cell.fill      = fill
            if col == rem_col:
                if remaining < 0:
                    cell.fill = PatternFill("solid", fgColor="FFE699")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="C00000")
                elif remaining == 0 and issued > 0:
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")
            if col in num_cols:
                cell.number_format = '#,##0'
        ws.row_dimensions[r_idx].height = 16
    ws.freeze_panes = f"A{hdr + 1}"
    last = hdr + len(results) + 1
    ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True, name="Calibri")
    for h in ["Stock", "Issued", "Remaining"]:
        if h in headers:
            ci = headers.index(h) + 1
            cl = get_column_letter(ci)
            cell = ws.cell(row=last, column=ci,
                           value=f"=SUM({cl}{hdr+1}:{cl}{last-1})")
            cell.font          = Font(bold=True, name="Calibri")
            cell.number_format = '#,##0'
            cell.fill          = PatternFill("solid", fgColor="D6E4F0")
            cell.border        = BORDER
    ws.row_dimensions[last].height = 18
    _xl_footer(ws, last, tc)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.oddHeader.center.text  = f"&B{APP_NAME}&B  |  {sheet_name}  |  &D  |  &T"
    ws.oddFooter.center.text  = FOOTER_TXT
    ws.print_title_rows       = "1:4"


def write_at_sheet(ws, agg_at, at_meta, source_files, now):
    headers    = ["SLoc","Material","Description","Batch","Issued"]
    col_widths = [8, 14, 42, 16, 12]
    tc = len(headers)
    _xl_info_header(ws, "@ Items", source_files, tc, now)
    hdr = 4
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hdr, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor="7030A0")
        cell.alignment = CTR_AL
        cell.border    = BORDER_HDR
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[hdr].height = 22
    FA = PatternFill("solid", fgColor="EAD1F5")
    FB = PatternFill("solid", fgColor="F7F0FC")
    for r_idx, (key, qty) in enumerate(sorted(agg_at.items()), hdr + 1):
        meta   = at_meta[key]
        fill   = FA if r_idx % 2 == 0 else FB
        values = [meta['sloc'], meta['material'], meta['description'], '', qty]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = LEFT_AL if col == 3 else CTR_AL
            cell.border    = BORDER
            cell.fill      = fill
            if col == 5: cell.number_format = '#,##0'
        ws.row_dimensions[r_idx].height = 16
    ws.freeze_panes = f"A{hdr + 1}"
    last = hdr + len(agg_at) + 1
    ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True, name="Calibri")
    cell = ws.cell(row=last, column=5, value=f"=SUM(E{hdr+1}:E{last-1})")
    cell.font = Font(bold=True, name="Calibri")
    cell.number_format = '#,##0'
    cell.fill   = PatternFill("solid", fgColor="E8D5F5")
    cell.border = BORDER
    ws.row_dimensions[last].height = 18
    _xl_footer(ws, last, tc)


def write_unmatched_sheet(ws, rows, sloc_name, source_files, now):
    headers    = ["SLoc","Material","Description","Batch","Issued"]
    col_widths = [8, 14, 42, 16, 12]
    tc = len(headers)
    _xl_info_header(ws, f"Unmatched — {sloc_name}", source_files, tc, now)
    hdr = 4
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hdr, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor="C00000")
        cell.alignment = CTR_AL
        cell.border    = BORDER_HDR
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[hdr].height = 22
    FA = PatternFill("solid", fgColor="FCE4D6")
    FB = PatternFill("solid", fgColor="FFF5F5")
    for r_idx, row in enumerate(rows, hdr + 1):
        fill   = FA if r_idx % 2 == 0 else FB
        values = [row['sloc'], row['material'], row['description'],
                  row['batch'], row['issued']]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = LEFT_AL if col == 3 else CTR_AL
            cell.border    = BORDER
            cell.fill      = fill
            if col == 5: cell.number_format = '#,##0'
        ws.row_dimensions[r_idx].height = 16
    ws.freeze_panes = f"A{hdr + 1}"
    last = hdr + len(rows) + 1
    ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True, name="Calibri")
    cell = ws.cell(row=last, column=5, value=f"=SUM(E{hdr+1}:E{last-1})")
    cell.font = Font(bold=True, name="Calibri")
    cell.number_format = '#,##0'
    cell.fill   = PatternFill("solid", fgColor="F2F2F2")
    cell.border = BORDER
    ws.row_dimensions[last].height = 18
    _xl_footer(ws, last, tc)


def build_excel_bytes(results_data, show_stock, show_issued, show_level):
    r   = results_data
    now = r['now']
    palette = ["DEEAF1","E2EFDA","FFF2CC","F2F2F2",
               "E8D5F5","FCE4D6","D9EAD3","CFE2F3"]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for i, s in enumerate(r['store_files']):
        write_stock_sheet(
            wb.create_sheet(s['sloc']),
            r['processed'][s['sloc']],
            palette[i % len(palette)],
            s['sloc'], r['source_files'], now,
            show_stock, show_issued, show_level)
    for sloc, rows in sorted(r['unmatched'].items()):
        write_unmatched_sheet(
            wb.create_sheet(f"{sloc}-UN"),
            rows, sloc, r['source_files'], now)
    if r['agg_at']:
        write_at_sheet(
            wb.create_sheet("@ Items"),
            r['agg_at'], r['at_meta'], r['source_files'], now)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  ٤. تصدير PDF — بدون تغيير
# ═══════════════════════════════════════════════════════════════════════════════

def _pdf_hf(canvas, doc, label, source_files, now):
    canvas.saveState()
    w, h = doc.pagesize
    canvas.setFillColor(colors.HexColor('#1F4E79'))
    canvas.rect(0, h - 1.4*cm, w, 1.4*cm, fill=1, stroke=0)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 11)
    canvas.drawString(1.5*cm, h - 0.9*cm, f"{APP_NAME}  —  {label}")
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(w - 1.5*cm, h - 0.9*cm,
                           f"{now.strftime('%Y-%m-%d')}   {now.strftime('%H:%M:%S')}")
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor('#AED6F1'))
    canvas.drawString(1.5*cm, h - 1.2*cm, f"Source: {', '.join(source_files)}")
    canvas.setFillColor(colors.HexColor('#1F4E79'))
    canvas.rect(0, 0, w, 0.8*cm, fill=1, stroke=0)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Oblique", 8)
    canvas.drawCentredString(w / 2, 0.25*cm, FOOTER_TXT)
    canvas.setFont("Helvetica", 7)
    canvas.drawRightString(w - 1.5*cm, 0.25*cm, f"Page {doc.page}")
    canvas.restoreState()


def build_pdf_bytes(label, rows_data, headers, col_widths_cm,
                    hdr_color, alt1, alt2, source_files, now, is_stock=False):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2.2*cm, bottomMargin=1.5*cm)
    styles  = getSampleStyleSheet()
    body_st = ParagraphStyle('body', parent=styles['Normal'],
                              fontSize=7.5, fontName='Helvetica', leading=10)
    desc_col   = 2
    table_data = [headers]
    for row in rows_data:
        table_data.append([
            Paragraph(str(v), body_st) if i == desc_col else v
            for i, v in enumerate(row)])
    t = Table(table_data,
              colWidths=[w * cm for w in col_widths_cm],
              repeatRows=1)
    rem_idx = headers.index("Remaining") + 1 if "Remaining" in headers else None
    cmds = [
        ('BACKGROUND',     (0,0),  (-1,0),  hdr_color),
        ('TEXTCOLOR',      (0,0),  (-1,0),  colors.white),
        ('FONTNAME',       (0,0),  (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0,0),  (-1,0),  8),
        ('ALIGN',          (0,0),  (-1,-1), 'CENTER'),
        ('ALIGN',          (2,1),  (2,-1),  'LEFT'),
        ('FONTNAME',       (0,1),  (-1,-1), 'Helvetica'),
        ('FONTSIZE',       (0,1),  (-1,-1), 7.5),
        ('ROWBACKGROUNDS', (0,1),  (-1,-1), [alt1, alt2]),
        ('GRID',           (0,0),  (-1,-1), 0.4, colors.HexColor('#B0C4D8')),
        ('VALIGN',         (0,0),  (-1,-1), 'MIDDLE'),
        ('TOPPADDING',     (0,0),  (-1,-1), 3),
        ('BOTTOMPADDING',  (0,0),  (-1,-1), 3),
    ]
    if is_stock and rem_idx:
        for i, row in enumerate(rows_data, 1):
            try:
                rem = float(str(row[rem_idx - 1]).replace(',', ''))
                if rem < 0:
                    cmds += [('BACKGROUND', (rem_idx-1,i), (rem_idx-1,i),
                               colors.HexColor('#FFE699')),
                              ('TEXTCOLOR',  (rem_idx-1,i), (rem_idx-1,i),
                               colors.HexColor('#C00000')),
                              ('FONTNAME',   (rem_idx-1,i), (rem_idx-1,i),
                               'Helvetica-Bold')]
                elif rem == 0:
                    cmds += [('BACKGROUND', (rem_idx-1,i), (rem_idx-1,i),
                               colors.HexColor('#FCE4D6'))]
            except:
                pass
    t.setStyle(TableStyle(cmds))
    doc.build([t],
              onFirstPage=lambda c, d: _pdf_hf(c, d, label, source_files, now),
              onLaterPages=lambda c, d: _pdf_hf(c, d, label, source_files, now))
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  ٥. واجهة Streamlit
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title=f"{APP_NAME} {APP_VER}",
    page_icon="▣",
    layout="wide"
)

st.markdown(f"""
<div style="background:#1F4E79;padding:14px 20px;border-radius:8px;margin-bottom:16px">
  <span style="color:white;font-size:20px;font-weight:bold">▣ {APP_NAME}</span>
  <span style="color:#AED6F1;font-size:13px;margin-left:10px">{APP_VER}</span>
</div>
""", unsafe_allow_html=True)

# ── الشريط الجانبي: المدخلات ─────────────────────────────────────────────────
with st.sidebar:
    st.header("① المدخلات")
    stock_files = st.file_uploader(
        "ملفات المخازن (HTM)",
        type=["htm","html"],
        accept_multiple_files=True,
        help="ارفع كل ملفات HTM الخاصة بالمخازن. اسم الملف يجب يبدأ بـ SLoc مثل: WH01-stock.htm"
    )
    issues_file = st.file_uploader(
        "ملف المنصرف (HTM)",
        type=["htm","html"],
        help="ملف MB51 الخاص بحركات الصرف"
    )

    st.divider()
    st.header("② خيارات الأعمدة")
    show_stock  = st.checkbox("Stock",       value=True)
    show_issued = st.checkbox("Issued",      value=True)
    show_level  = st.checkbox("Match Level", value=True)

    st.divider()
    st.caption(f"Remaining و SLoc/Material/Batch/Description ثوابت دائماً")

# ── الزر الرئيسي ─────────────────────────────────────────────────────────────
col_run, col_info = st.columns([1, 3])
with col_run:
    run_btn = st.button("▶ تشغيل المعالجة", type="primary",
                        disabled=(not stock_files or not issues_file))

if not stock_files:
    st.info("ارفع ملفات المخازن من الشريط الجانبي للبدء")
    st.stop()
if not issues_file:
    st.info("ارفع ملف المنصرف من الشريط الجانبي للبدء")
    st.stop()

# ── تشغيل المعالجة ────────────────────────────────────────────────────────────
if run_btn or st.session_state.get('processed'):

    if run_btn:
        with st.spinner("جارٍ المعالجة…"):

            # تحديد SLoc من اسم الملف
            store_files = []
            for f in stock_files:
                if f.name.lower().startswith('good'):
                    continue
                sloc = f.name.split('-')[0].strip()
                store_files.append({'sloc': sloc, 'filename': f.name,
                                    'bytes': f.read()})
            store_files.sort(key=lambda x: x['sloc'])

            # قراءة Issues
            issues_bytes = issues_file.read()
            issues_agg, agg_at, at_meta = parse_issues_bytes(issues_bytes)

            # معالجة كل مخزن
            known_slocs = {s['sloc'] for s in store_files}
            processed   = {}
            for s in store_files:
                stock = parse_stock_bytes(s['bytes'], s['sloc'])
                processed[s['sloc']] = process(stock, issues_agg)

            unmatched    = get_unmatched(issues_agg, known_slocs)
            source_files = ([issues_file.name] +
                            [s['filename'] for s in store_files])

            st.session_state['processed'] = {
                'store_files':  store_files,
                'processed':    processed,
                'unmatched':    unmatched,
                'agg_at':       agg_at,
                'at_meta':      at_meta,
                'source_files': source_files,
                'now':          datetime.datetime.now(),
            }

    r = st.session_state.get('processed')
    if not r:
        st.stop()

    # ── ملخص النتائج ─────────────────────────────────────────────────────────
    st.subheader("③ ملخص النتائج")
    summary_rows = []
    for s in r['store_files']:
        sloc    = s['sloc']
        results = r['processed'][sloc]
        matched = sum(1 for rv in results if rv['issued'] > 0)
        summary_rows.append({
            "المخزن": sloc,
            "إجمالي الأصناف": len(results),
            "مخصوم منه": matched,
            "بدون خصم": len(results) - matched,
            "النوع": "مخزن"
        })
    for sloc, rows in sorted(r['unmatched'].items()):
        summary_rows.append({
            "المخزن": sloc,
            "إجمالي الأصناف": len(rows),
            "مخصوم منه": len(rows),
            "بدون خصم": 0,
            "النوع": "⚠ Unmatched"
        })
    if r['agg_at']:
        summary_rows.append({
            "المخزن": "@ Items",
            "إجمالي الأصناف": len(r['agg_at']),
            "مخصوم منه": len(r['agg_at']),
            "بدون خصم": 0,
            "النوع": "◈ @ Items"
        })
    st.dataframe(summary_rows, use_container_width=True, hide_index=True)

    # ── تصدير Excel ──────────────────────────────────────────────────────────
    st.subheader("④ تصدير Excel")
    xl_bytes = build_excel_bytes(r, show_stock, show_issued, show_level)
    fname_xl = f"NEAM_Stock_{r['now'].strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        label="⬇ تحميل Excel",
        data=xl_bytes,
        file_name=fname_xl,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ── تصدير PDF ────────────────────────────────────────────────────────────
    st.subheader("⑤ تصدير PDF")
    st.caption("اختر الشيتات التي تريد تصديرها — كل شيت ينزل كملف منفصل")

    hdr_map = {'stock':     colors.HexColor('#1F4E79'),
               'at':        colors.HexColor('#4B1878'),
               'unmatched': colors.HexColor('#8B0000')}
    alt_map = {
        'stock':     (colors.HexColor('#DEEAF1'), colors.HexColor('#F0F7FF')),
        'at':        (colors.HexColor('#EAD1F5'), colors.HexColor('#F7F0FC')),
        'unmatched': (colors.HexColor('#FCE4D6'), colors.HexColor('#FFF5F5')),
    }

    all_sheets = []
    for s in r['store_files']:
        all_sheets.append({'label': s['sloc'], 'type': 'stock'})
    for sloc in sorted(r['unmatched'].keys()):
        all_sheets.append({'label': f"Unmatched-{sloc}", 'type': 'unmatched'})
    if r['agg_at']:
        all_sheets.append({'label': '@ Items', 'type': 'at'})

    pdf_cols = st.columns(min(len(all_sheets), 4))
    pdf_sel  = {}
    for i, info in enumerate(all_sheets):
        with pdf_cols[i % 4]:
            icon = {"stock":"▣","unmatched":"⚠","at":"◈"}.get(info['type'],"■")
            pdf_sel[info['label']] = st.checkbox(
                f"{icon} {info['label']}", value=True,
                key=f"pdf_{info['label']}")

    for info in all_sheets:
        label, typ = info['label'], info['type']
        if not pdf_sel.get(label):
            continue

        if typ == 'stock':
            results_s       = r['processed'].get(label, [])
            headers, widths = filter_headers(show_stock, show_issued, show_level)
            cw_map = {"SLoc":1.2,"Material":2.0,"Description":6.0,
                      "Batch":2.2,"Stock":1.8,"Issued":1.8,
                      "Remaining":2.0,"Match Level":1.8}
            cw       = [cw_map.get(h, 2.0) for h in headers]
            pdf_rows = []
            for rv in results_s:
                vals = row_to_values(rv, show_stock, show_issued, show_level)
                pdf_rows.append([
                    f"{v:,.0f}" if isinstance(v, (int, float)) else v
                    for v in vals])
        elif typ == 'unmatched':
            sloc_name = label.replace("Unmatched-", "")
            headers   = ["SLoc","Material","Description","Batch","Issued"]
            cw        = [1.2, 2.0, 8.8, 2.2, 2.0]
            pdf_rows  = [[rv['sloc'], rv['material'],
                          rv['description'][:50],
                          rv['batch'], f"{rv['issued']:,.0f}"]
                         for rv in r['unmatched'].get(sloc_name, [])]
        elif typ == 'at':
            headers  = ["SLoc","Material","Description","Issued"]
            cw       = [1.2, 2.0, 11.0, 2.0]
            pdf_rows = [[r['at_meta'][k]['sloc'],
                         r['at_meta'][k]['material'],
                         r['at_meta'][k]['description'][:55],
                         f"{v:,.0f}"]
                        for k, v in sorted(r['agg_at'].items())]
        else:
            continue

        pdf_bytes = build_pdf_bytes(
            label, pdf_rows, headers, cw,
            hdr_map.get(typ, colors.HexColor('#1F4E79')),
            alt_map.get(typ, (colors.white,)*2)[0],
            alt_map.get(typ, (colors.white,)*2)[1],
            r['source_files'], r['now'],
            is_stock=(typ == 'stock'))

        fname_pdf = f"NEAM_{label.replace(' ','_')}_{r['now'].strftime('%Y%m%d_%H%M%S')}.pdf"
        st.download_button(
            label=f"⬇ PDF — {label}",
            data=pdf_bytes,
            file_name=fname_pdf,
            mime="application/pdf",
            key=f"dl_pdf_{label}"
        )

    st.divider()
    st.caption(f"▣ {FOOTER_TXT}  |  {APP_NAME} {APP_VER}")
